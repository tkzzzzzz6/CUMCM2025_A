# -*- coding: utf-8 -*-
"""
Problem 2: 三枚烟幕弹对单导弹的协同遮蔽优化
依赖: numpy, pandas, openpyxl
- [最终版 - 格式化输出] 保持核心优化逻辑不变。
- 按照用户提供的图片模板，格式化控制台输出和生成的result1.xlsx文件。
"""
import math
import numpy as np
import time
import random
import copy
import pandas as pd

# -----------------------------
# 1) 题设常量与几何/运动模型
# -----------------------------
g = 9.80665
R = 10.0
cloud_sink = 3.0
M0 = np.array([20000.0, 0.0, 2000.0], dtype=float)
to_origin = -M0 / np.linalg.norm(M0)
vM = 300.0 * to_origin
U0 = np.array([17800.0, 0.0, 1800.0], dtype=float)
T_center = np.array([0.0, 200.0, 5.0], dtype=float)

def M_pos(t: float) -> np.ndarray:
    return M0 + vM * t

# -------------------------------------
# 2) 核心计算与适应度函数
# -------------------------------------
def point_to_segment_distance(c: np.ndarray, a: np.ndarray, b: np.ndarray):
    ab = b - a
    denom = float(np.dot(ab, ab))
    if denom < 1e-9:
        return np.linalg.norm(c - a), 0.0
    lam = float(np.dot(c - a, ab) / denom)
    d = float(np.linalg.norm(c - (a + max(0.0, min(1.0, lam)) * ab)))
    return d, lam

def calculate_total_coverage_union(params, N_grid=1001, is_final_calc=False):
    angle, speed, t_r1, t_f1, dt_r2, t_f2, dt_r3, t_f3 = params
    t_releases = [t_r1, t_r1 + dt_r2, t_r1 + dt_r2 + dt_r3]
    t_fuzes = [t_f1, t_f2, t_f3]
    angle_rad = np.deg2rad(angle)
    uav_dir = np.array([np.cos(angle_rad), np.sin(angle_rad), 0.0])

    bombs_data = []
    for t_r, t_f in zip(t_releases, t_fuzes):
        p_release = U0 + uav_dir * speed * t_r
        explosion_point = np.array([
            p_release[0] + uav_dir[0] * speed * t_f,
            p_release[1] + uav_dir[1] * speed * t_f,
            U0[2] - 0.5 * g * (t_f ** 2)
        ])
        t_explode = t_r + t_f
        bombs_data.append({'E': explosion_point, 't_exp': t_explode, 'P_rel': p_release})

    min_exp_time = min(b['t_exp'] for b in bombs_data)
    max_exp_time = max(b['t_exp'] for b in bombs_data)
    t0, t1 = min_exp_time, max_exp_time + 20.0
    t_grid = np.linspace(t0, t1, N_grid)
    dt = t_grid[1] - t_grid[0]
    
    union_mask = np.zeros_like(t_grid, dtype=bool)
    for i, t in enumerate(t_grid):
        m_pos = M_pos(t)
        is_obscured_this_step = False
        for bomb in bombs_data:
            if t >= bomb['t_exp']:
                c_pos = bomb['E'] - np.array([0, 0, cloud_sink * (t - bomb['t_exp'])])
                d, lam = point_to_segment_distance(c_pos, m_pos, T_center)
                if d <= R and 0.0 <= lam <= 1.0:
                    is_obscured_this_step = True
                    break
        union_mask[i] = is_obscured_this_step
        
    total_duration = np.sum(union_mask) * dt
    
    if is_final_calc:
        return total_duration, bombs_data
    else:
        return total_duration

# ---------------------------------------------
# 3) 遗传算法 (Genetic Algorithm) 模块
# ---------------------------------------------
POP_SIZE = 100
GENERATIONS = 150
CXPB, MUTPB = 0.9, 0.2
ELITE_COUNT = 2
ETA = 20
PARAM_BOUNDS = [(0, 360),(70, 140),(0.01, 8.0),(0.5, 8.0),(1.0, 8.0),(0.5, 8.0),(1.0, 8.0),(0.5, 8.0)]

def check_bounds(val, bounds): return max(bounds[0], min(bounds[1], val))
def create_individual(): return [random.uniform(b[0], b[1]) for b in PARAM_BOUNDS]
def crossover_sbx(p1,p2): c1,c2=copy.deepcopy(p1),copy.deepcopy(p2); [crossover_sbx_gene(c1,c2,i) for i in range(len(p1))]; return c1,c2
def crossover_sbx_gene(c1,c2,i):
    if random.random()>0.5: return
    y1,y2,(yl,yu)=min(c1[i],c2[i]),max(c1[i],c2[i]),PARAM_BOUNDS[i]; u=random.random()
    if y2-y1>1e-6:
        beta=1.0+(2.0*(y1-yl)/(y2-y1)); alpha=2.0-beta**(-(ETA+1.0)); beta_q=(u*alpha)**(1.0/(ETA+1.0)) if u<=1.0/alpha else (1.0/(2.0-u*alpha))**(1.0/(ETA+1.0)); c1[i]=0.5*((y1+y2)-beta_q*(y2-y1))
        beta=1.0+(2.0*(yu-y2)/(y2-y1)); alpha=2.0-beta**(-(ETA+1.0)); beta_q=(u*alpha)**(1.0/(ETA+1.0)) if u<=1.0/alpha else (1.0/(2.0-u*alpha))**(1.0/(ETA+1.0)); c2[i]=0.5*((y1+y2)+beta_q*(y2-y1))
    c1[i],c2[i]=check_bounds(c1[i],PARAM_BOUNDS[i]),check_bounds(c2[i],PARAM_BOUNDS[i])
def mutate_polynomial(ind): [mutate_polynomial_gene(ind,i) for i in range(len(ind))]; return ind
def mutate_polynomial_gene(ind,i):
    if random.random()<MUTPB:
        yl,yu=PARAM_BOUNDS[i]; delta1,delta2=(ind[i]-yl)/(yu-yl),(yu-ind[i])/(yu-yl); u,mut_pow=random.random(),1.0/(ETA+1.0)
        if u<0.5: val=2.0*u+(1.0-2.0*u)*(1.0-delta1)**(ETA+1.0); delta_q=val**mut_pow-1.0
        else: val=2.0*(1.0-u)+2.0*(u-0.5)*(1.0-delta2)**(ETA+1.0); delta_q=1.0-val**mut_pow
        ind[i]+=delta_q*(yu-yl); ind[i]=check_bounds(ind[i],PARAM_BOUNDS[i])

def run_ga_optimization():
    print("=== 开始执行三枚烟幕弹协同策略优化 ===")
    population = [create_individual() for _ in range(POP_SIZE)]
    best_ind_overall, best_fitness_overall = None, -1.0
    start_time = time.time()

    for gen in range(GENERATIONS):
        fitnesses = [calculate_total_coverage_union(ind) for ind in population]
        elite_indices = sorted(range(len(fitnesses)), key=lambda k: fitnesses[k], reverse=True)[:ELITE_COUNT]
        elites = [population[i] for i in elite_indices]

        if fitnesses[elite_indices[0]] > best_fitness_overall:
            best_fitness_overall = fitnesses[elite_indices[0]]
            best_ind_overall = copy.deepcopy(elites[0])

        print(f"代数 {gen+1:3d}/{GENERATIONS} | "f"本代最长: {max(fitnesses):.3f} s | "f"历史最长: {best_fitness_overall:.3f} s")

        offspring = elites
        while len(offspring) < POP_SIZE:
            p1_idx,p2_idx=random.sample(range(POP_SIZE),2); p1=population[p1_idx] if fitnesses[p1_idx]>fitnesses[p2_idx] else population[p2_idx]
            p3_idx,p4_idx=random.sample(range(POP_SIZE),2); p2=population[p3_idx] if fitnesses[p3_idx]>fitnesses[p4_idx] else population[p3_idx]
            if random.random()<CXPB: c1,c2=crossover_sbx(p1,p2)
            else: c1,c2=p1,p2
            offspring.append(mutate_polynomial(c1));
            if len(offspring)<POP_SIZE: offspring.append(mutate_polynomial(c2))
        population = offspring

    print(f"\n优化完成，耗时: {time.time() - start_time:.2f} s")
    return best_ind_overall

# -----------------------------
# 4) 主流程与格式化输出
# -----------------------------
def format_and_output_results(best_params):
    """
    对最优策略进行高精度计算，并按指定格式输出到控制台和Excel文件。
    """
    print("\n=== 对最优策略进行高精度计算并生成报告... ===")
    
    # 1. 高精度计算
    total_duration_precise, bombs_data_precise = calculate_total_coverage_union(
        best_params, N_grid=5001, is_final_calc=True
    )
    
    # 2. 准备数据
    angle, speed, t_r1, t_f1, dt_r2, t_f2, dt_r3, t_f3 = best_params
    t_releases = [t_r1, t_r1 + dt_r2, t_r1 + dt_r2 + dt_r3]
    t_fuzes = [t_f1, t_f2, t_f3]

    # 3. 构建DataFrame
    # 定义多级列标题
    headers = [
        ('无人机信息', '飞行速度(m/s)'), ('无人机信息', '飞行方向(度)'),
        ('烟幕弹1', '投放时间(s)'), ('烟幕弹1', '引信时间(s)'),
        ('烟幕弹1', '投放点x'), ('烟幕弹1', '投放点y'), ('烟幕弹1', '投放点z'),
        ('烟幕弹1', '起爆点x'), ('烟幕弹1', '起爆点y'), ('烟幕弹1', '起爆点z'),
        ('烟幕弹2', '投放时间(s)'), ('烟幕弹2', '引信时间(s)'),
        ('烟幕弹2', '投放点x'), ('烟幕弹2', '投放点y'), ('烟幕弹2', '投放点z'),
        ('烟幕弹2', '起爆点x'), ('烟幕弹2', '起爆点y'), ('烟幕弹2', '起爆点z'),
        ('烟幕弹3', '投放时间(s)'), ('烟幕弹3', '引信时间(s)'),
        ('烟幕弹3', '投放点x'), ('烟幕弹3', '投放点y'), ('烟幕弹3', '投放点z'),
        ('烟幕弹3', '起爆点x'), ('烟幕弹3', '起爆点y'), ('烟幕弹3', '起爆点z'),
    ]
    columns = pd.MultiIndex.from_tuples(headers)
    
    # 创建数据行
    data_row = []
    data_row.extend([speed, angle])
    for i in range(3):
        p_rel = bombs_data_precise[i]['P_rel']
        bomb_e = bombs_data_precise[i]['E']
        data_row.extend([
            t_releases[i], t_fuzes[i],
            p_rel[0], p_rel[1], p_rel[2],
            bomb_e[0], bomb_e[1], bomb_e[2]
        ])
    
    df = pd.DataFrame([data_row], columns=columns)

    # 4. 打印到控制台
    print("\n" + "="*80)
    print(" " * 28 + "无人机FY1 烟幕干扰弹投放策略")
    print("="*80)
    print(f" 无人机信息 ".center(24, '-'))
    print(f"  飞行速度: {speed:.2f} m/s")
    print(f"  飞行方向: {angle:.2f} 度")
    print("-"*24)
    for i in range(3):
        print(f" 烟幕弹{i+1} ".center(50, '-'))
        print(f"  投放时间: {t_releases[i]:.2f} s".ljust(25) + f"引信时间: {t_fuzes[i]:.2f} s")
        p = bombs_data_precise[i]['P_rel']
        e = bombs_data_precise[i]['E']
        print(f"  投放点 (x,y,z): ({p[0]:.1f}, {p[1]:.1f}, {p[2]:.1f})")
        print(f"  起爆点 (x,y,z): ({e[0]:.1f}, {e[1]:.1f}, {e[2]:.1f})")
    print("="*80)
    print(f"总有效遮蔽时长: {total_duration_precise:.4f} s")
    print("="*80)

    # 5. 保存到Excel文件
    output_filename = 'result1.xlsx'
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        workbook = writer.book
        sheet_name = '投放策略'
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name] # 删除默认创建的空sheet
        
        # 写入标题
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.merge_cells('A1:Z1')
        title_cell = worksheet['A1']
        title_cell.value = '无人机FY1 烟幕干扰弹投放策略'
        from openpyxl.styles import Alignment, Font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(bold=True, size=16)
        worksheet.row_dimensions[1].height = 30
        
        # 写入DataFrame
        df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
        
        # 写入总时长
        summary_row = df.shape[0] + 5 # DataFrame写入后隔一行
        worksheet.cell(row=summary_row, column=1, value='总有效遮蔽时长(s)').font = Font(bold=True)
        worksheet.cell(row=summary_row, column=2, value=f"{total_duration_precise:.4f}")

        # 调整格式
        # 合并一级标题单元格
        worksheet.merge_cells('A3:B3'); worksheet['A3'].value = '无人机信息'; worksheet['A3'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('C3:J3'); worksheet['C3'].value = '烟幕弹1'; worksheet['C3'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('K3:R3'); worksheet['K3'].value = '烟幕弹2'; worksheet['K3'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('S3:Z3'); worksheet['S3'].value = '烟幕弹3'; worksheet['S3'].alignment = Alignment(horizontal='center')
        
        # 调整列宽
        for col_idx, col in enumerate(worksheet.columns, 1):
            worksheet.column_dimensions[col[0].column_letter].width = 15

    print(f"\n详细策略已保存到文件: {output_filename}")

if __name__ == "__main__":
    best_params = run_ga_optimization()
    if best_params:
        format_and_output_results(best_params)