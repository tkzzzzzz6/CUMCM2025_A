# -*- coding: utf-8 -*-
"""
Problem 5: 五架无人机，每架至多投放三枚烟幕弹，对三枚导弹 M1/M2/M3 的协同遮蔽优化（PSO）
依赖: numpy, matplotlib, openpyxl（可选，若无则跳过 Excel 写入）

输出:
- code/p5/output_PSO/summary_p5_PSO.txt
- code/p5/output_PSO/pso_progress.png
- 附件/result3.xlsx 的工作表 'Q5'
"""
import os
import time
import math
import random
import numpy as np
import matplotlib.pyplot as plt

try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None
    Workbook = None

# Matplotlib 中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# -----------------------------
# 1) 物理参数与初始条件
# -----------------------------
g = 9.80665
SMOKE_EFFECTIVE_RADIUS = 10.0
SINK_SPEED = 3.0
EFFECTIVE_TIME = 20.0
MISSILE_SPEED = 300.0

# 五架无人机初始位置（FY1..FY5）
UAV_STARTS = [
    np.array([17800.0,     0.0, 1800.0], dtype=float),  # FY1
    np.array([12000.0,  1400.0, 1400.0], dtype=float),  # FY2
    np.array([ 6000.0, -3000.0,  700.0], dtype=float),  # FY3
    np.array([11000.0,  2000.0, 1800.0], dtype=float),  # FY4
    np.array([13000.0, -2000.0, 1300.0], dtype=float),  # FY5
]

# 导弹与目标
MISSILE_START_M1 = np.array([20000.0,     0.0, 2000.0], dtype=float)
MISSILE_START_M2 = np.array([19000.0,   600.0, 2100.0], dtype=float)
MISSILE_START_M3 = np.array([18000.0,  -600.0, 1900.0], dtype=float)
TARGET_CENTER = np.array([0.0, 200.0, 5.0], dtype=float)

R = SMOKE_EFFECTIVE_RADIUS
cloud_sink = SINK_SPEED
M0_list = [MISSILE_START_M1, MISSILE_START_M2, MISSILE_START_M3]
vM_list = [MISSILE_SPEED * (-M0 / np.linalg.norm(M0)) for M0 in M0_list]
T_center = TARGET_CENTER

def M_pos_j(t: float, j: int) -> np.ndarray:
    M0 = M0_list[j]
    vM = vM_list[j]
    return M0 + vM * t

# -----------------------------
# 2) 几何与适应度
# -----------------------------
def point_to_segment_distance(c: np.ndarray, a: np.ndarray, b: np.ndarray):
    ab = b - a
    denom = float(np.dot(ab, ab))
    if denom < 1e-9:
        return np.linalg.norm(c - a), 0.0
    lam = float(np.dot(c - a, ab) / denom)
    lam_clamped = max(0.0, min(1.0, lam))
    d = float(np.linalg.norm(c - (a + lam_clamped * ab)))
    return d, lam_clamped

NUM_UAV = 5
BOMBS_PER_UAV = 3
NUM_MISSILES = 3

def _unpack_params(params):
    """
    连续变量部分（前 8*NUM_UAV = 40 个）:
      对每个 UAV i: [angle_i, speed_i, t_r1_i, dt_r2_i, dt_r3_i, t_f1_i, t_f2_i, t_f3_i]
    指派变量部分（后 15 个）:
      对所有 15 枚弹（按 FY1 的 3 枚，FY2 的 3 枚 ... FY5 的 3 枚次序）: s_b ∈ [0,3) → m_idx = int(s_b)∈{0,1,2}
    返回:
      uavs: 列表，元素为 (angle, speed, t_r1, dt_r2, dt_r3, [t_f1,t_f2,t_f3])
      assign: 长度 15 的整型列表，每个值 ∈ {0,1,2}
    """
    uavs = []
    for i in range(NUM_UAV):
        base = 8 * i
        angle = params[base + 0]
        speed = params[base + 1]
        t_r1  = params[base + 2]
        dt_r2 = params[base + 3]
        dt_r3 = params[base + 4]
        t_f1  = params[base + 5]
        t_f2  = params[base + 6]
        t_f3  = params[base + 7]
        uavs.append((angle, speed, t_r1, dt_r2, dt_r3, [t_f1, t_f2, t_f3]))

    assign = []
    start_idx = 8 * NUM_UAV
    for b in range(NUM_UAV * BOMBS_PER_UAV):
        s = params[start_idx + b]
        m_idx = int(max(0, min(2.999, s)))  # 0,1,2
        assign.append(m_idx)
    return uavs, assign

def calculate_total_coverage_union(params, N_grid=1001, is_final_calc=False):
    # 钳制边界
    clamped = [check_bounds(v, PARAM_BOUNDS[i]) for i, v in enumerate(params)]
    uavs, assign = _unpack_params(clamped)

    bombs_data = []  # 记录 15 枚弹
    # 构建每枚弹的几何信息
    for i in range(NUM_UAV):
        angle, speed, t_r1, dt_r2, dt_r3, t_f_list = uavs[i]
        angle_rad = np.deg2rad(angle % 360.0)
        uav_dir = np.array([np.cos(angle_rad), np.sin(angle_rad), 0.0])
        U0 = UAV_STARTS[i]

        t_rels = [t_r1, t_r1 + dt_r2, t_r1 + dt_r3 + dt_r2]  # 三次释放时刻（相邻≥1由边界保证）
        for k in range(BOMBS_PER_UAV):
            t_rel = t_rels[k]
            t_fuz = t_f_list[k]
            p_release = U0 + uav_dir * speed * t_rel
            explosion_point = np.array([
                p_release[0] + uav_dir[0] * speed * t_fuz,
                p_release[1] + uav_dir[1] * speed * t_fuz,
                U0[2] - 0.5 * g * (t_fuz ** 2)
            ])
            t_explode = t_rel + t_fuz
            global_idx = i * BOMBS_PER_UAV + k
            bombs_data.append({
                'E': explosion_point,
                't_exp': t_explode,
                'P_rel': p_release,
                'uav_idx': i,
                'bomb_idx': k,
                'm_idx': assign[global_idx]
            })

    # 时间窗
    min_exp_time = min(b['t_exp'] for b in bombs_data)
    t0 = min_exp_time
    t1 = max(b['t_exp'] + EFFECTIVE_TIME for b in bombs_data)
    t_grid = np.linspace(t0, t1, N_grid)
    dt = t_grid[1] - t_grid[0]

    # 三个导弹的并集时长
    union_masks = [np.zeros_like(t_grid, dtype=bool) for _ in range(NUM_MISSILES)]
    for gi, t in enumerate(t_grid):
        for j in range(NUM_MISSILES):
            m_pos = M_pos_j(t, j)
            covered = False
            for b in bombs_data:
                if b['m_idx'] != j:
                    continue
                if t >= b['t_exp'] and t <= b['t_exp'] + EFFECTIVE_TIME:
                    c_pos = b['E'] - np.array([0, 0, cloud_sink * (t - b['t_exp'])])
                    d, _ = point_to_segment_distance(c_pos, m_pos, T_center)
                    if d <= R:
                        covered = True
                        break
            union_masks[j][gi] = covered

    durations = [float(np.sum(mask) * dt) for mask in union_masks]
    total_duration = float(sum(durations))
    if is_final_calc:
        return total_duration, bombs_data, durations
    return total_duration

# 缓存
FITNESS_CACHE = {}

def get_fitness(params, fidelity='low'):
    key = (tuple([round(check_bounds(v, PARAM_BOUNDS[i]), 3) for i, v in enumerate(params)]), fidelity)
    if key in FITNESS_CACHE:
        return FITNESS_CACHE[key]
    N = 301 if fidelity == 'low' else 1501
    val = calculate_total_coverage_union(params, N_grid=N, is_final_calc=False)
    FITNESS_CACHE[key] = val
    return val

# -----------------------------
# 3) PSO
# -----------------------------
# 每个 UAV: [angle, speed, t_r1, dt_r2, dt_r3, t_f1, t_f2, t_f3] 共 8 个
# 指派变量: 15 个，范围 [0, 2.999]，在评估时取 int → {0,1,2}
PARAM_BOUNDS = []
for _ in range(NUM_UAV):
    PARAM_BOUNDS.extend([
        (0.0, 360.0),      # angle
        (70.0, 140.0),     # speed
        (0.0, 20.0),       # t_r1
        (1.0, 15.0),       # dt_r2 >= 1
        (1.0, 15.0),       # dt_r3 >= 1
        (0.005, 12.0),     # t_f1
        (0.005, 12.0),     # t_f2
        (0.005, 12.0),     # t_f3
    ])
for _ in range(NUM_UAV * BOMBS_PER_UAV):
    PARAM_BOUNDS.append((0.0, 2.999))  # assignment real-coded

def check_bounds(val, bounds):
    return max(bounds[0], min(bounds[1], float(val)))

def project_params(params):
    q = [check_bounds(v, b) for v, b in zip(params, PARAM_BOUNDS)]
    # 角度归一：每 8 个为一组，角度在组首
    for i in range(NUM_UAV):
        idx = 8*i
        q[idx] = q[idx] % 360.0
    return q

def run_pso_optimization():
    print("=== 开始执行 五机至多三弹协同策略优化 (PSO, Q5) ===")
    random.seed(42); np.random.seed(42)

    SWARM_SIZE = 120
    ITERATIONS = 10000
    W = 0.6
    C1 = 1.6
    C2 = 1.6

    dim = len(PARAM_BOUNDS)
    positions = np.array([[random.uniform(b[0], b[1]) for b in PARAM_BOUNDS] for _ in range(SWARM_SIZE)], dtype=float)
    velocities = np.zeros((SWARM_SIZE, dim), dtype=float)

    pbest = positions.copy()
    pbest_fit = np.array([get_fitness(project_params(list(p)), fidelity='low') for p in pbest])
    g_idx = int(np.argmax(pbest_fit))
    gbest = pbest[g_idx].copy()
    gbest_fit = float(pbest_fit[g_idx])

    start_time = time.time()
    gen_best_list = []

    for it in range(ITERATIONS):
        r1 = np.random.rand(SWARM_SIZE, dim)
        r2 = np.random.rand(SWARM_SIZE, dim)
        velocities = W*velocities + C1*r1*(pbest - positions) + C2*r2*(gbest - positions)
        positions = positions + velocities
        positions = np.array([project_params(list(p)) for p in positions], dtype=float)

        fits = np.array([get_fitness(list(p), fidelity='low') for p in positions])
        gen_best_list.append(float(np.max(fits)))

        improved = fits > pbest_fit
        pbest[improved] = positions[improved]
        pbest_fit[improved] = fits[improved]

        g_idx = int(np.argmax(pbest_fit))
        gbest_candidate = pbest[g_idx].copy()
        gbest_candidate_fit = get_fitness(list(gbest_candidate), fidelity='high')
        if gbest_candidate_fit > gbest_fit:
            gbest, gbest_fit = gbest_candidate, gbest_candidate_fit
            print("[刷新历史遮挡最长] 当前最佳(高保真): {:.3f} s".format(gbest_fit))

        print(f"迭代 {it+1:3d}/{ITERATIONS} | 本轮最佳(低保真): {np.max(fits):.3f} s | 历史最佳(高保真): {gbest_fit:.3f} s")

    # 输出摘要与进度图
    out_dir = os.path.join(os.path.dirname(__file__), 'output_PSO')
    os.makedirs(out_dir, exist_ok=True)
    try:
        with open(os.path.join(out_dir, 'summary_p5_PSO.txt'), 'w', encoding='utf-8') as f:
            f.write("[总体最优策略(高保真评估, Q5)]\n")
            f.write("最长有效遮蔽并集时长(三导弹求和): {:.3f} s\n".format(gbest_fit))
            # 计算一次高保真，拿到分导弹时长
            total_fit_precise, bombs_data_precise, durations = calculate_total_coverage_union(list(gbest), N_grid=3001, is_final_calc=True)
            f.write("分导弹并集时长: M1={:.3f}s, M2={:.3f}s, M3={:.3f}s\n".format(durations[0], durations[1], durations[2]))
            uavs, assign = _unpack_params(gbest)
            for i, (angle, speed, t_r1, dt_r2, dt_r3, t_f_list) in enumerate(uavs):
                f.write(f"FY{i+1}: 角度 {angle:.3f} 度, 速度 {speed:.2f} m/s, t_r1 {t_r1:.3f} s, dt2 {dt_r2:.3f} s, dt3 {dt_r3:.3f} s, t_f {t_f_list}\n")
    except Exception as e:
        print("写入摘要失败:", e)

    try:
        plt.figure(figsize=(10,4.5))
        x = list(range(1, len(gen_best_list)+1))
        plt.plot(x, gen_best_list, 'o-', linewidth=1.8, markersize=4, color='#1f77b4', alpha=0.9, label='本代最长', zorder=2)
        plt.xlabel('代数'); plt.ylabel('遮蔽时间总和 (s)'); plt.title('PSO 优化进度 (Q5)')
        plt.legend(); plt.grid(alpha=0.3); plt.tight_layout()
        plt.savefig(os.path.join(out_dir, 'pso_progress.png'), dpi=150)
        plt.close()
    except Exception as e:
        print("保存进度图失败:", e)

    return list(gbest), gbest_fit

# -----------------------------
# 4) 结果输出（控制台与 Excel）
# -----------------------------
def calculate_individual_durations_for_missile(bombs_data, missile_idx, N_grid=4001):
    min_exp_time = min(b['t_exp'] for b in bombs_data)
    t1 = max(b['t_exp'] + EFFECTIVE_TIME for b in bombs_data)
    t_grid = np.linspace(min_exp_time, t1, N_grid)
    dt = t_grid[1] - t_grid[0]
    durations = []
    for bomb in bombs_data:
        if bomb['m_idx'] != missile_idx:
            durations.append(0.0)
            continue
        mask = np.zeros_like(t_grid, dtype=bool)
        for i, t in enumerate(t_grid):
            if t >= bomb['t_exp'] and t <= bomb['t_exp'] + EFFECTIVE_TIME:
                c_pos = bomb['E'] - np.array([0, 0, cloud_sink * (t - bomb['t_exp'])])
                d, _ = point_to_segment_distance(c_pos, M_pos_j(t, missile_idx), T_center)
                if d <= R:
                    mask[i] = True
        durations.append(float(np.sum(mask) * dt))
    return durations

def format_and_output_results(best_params):
    print("\n=== 对最优策略进行高精度计算并生成报告 (Q5) ===")
    total_duration_precise, bombs_data_precise, per_missile = calculate_total_coverage_union(
        best_params, N_grid=5001, is_final_calc=True
    )
    uavs, assign = _unpack_params(best_params)

    print("总有效遮蔽并集时长总和 (s): {:.4f}".format(total_duration_precise))
    print("分导弹并集时长: M1={:.3f}s, M2={:.3f}s, M3={:.3f}s".format(per_missile[0], per_missile[1], per_missile[2]))
    for idx, bomb in enumerate(bombs_data_precise):
        pr = bomb['P_rel']; e = bomb['E']
        uav_idx = bomb['uav_idx']
        angle, speed, t_r1, dt_r2, dt_r3, t_f_list = uavs[uav_idx]
        # 还原该弹的 t_rel 与 t_fuz
        if bomb['bomb_idx'] == 0:
            t_rel = t_r1; t_fuz = t_f_list[0]
        elif bomb['bomb_idx'] == 1:
            t_rel = t_r1 + dt_r2; t_fuz = t_f_list[1]
        else:
            t_rel = t_r1 + dt_r2 + dt_r3; t_fuz = t_f_list[2]
        print(f"FY{uav_idx+1}-B{bomb['bomb_idx']+1} -> M{bomb['m_idx']+1}")
        print("  航向角(度): {:.2f}".format(angle))
        print("  速度(m/s): {:.2f}".format(speed))
        print("  投放时间(s): {:.3f}".format(t_rel))
        print("  引信时间(s): {:.3f}".format(t_fuz))
        print("  投放点(m): ({:.3f}, {:.3f}, {:.3f})".format(pr[0], pr[1], pr[2]))
        print("  起爆点(m): ({:.3f}, {:.3f}, {:.3f})".format(e[0], e[1], e[2]))

    # 写入 Excel: 附件/result3.xlsx
    if Workbook is None:
        print("未安装 openpyxl，跳过 Excel 写入。请在 requirements.txt 中包含 openpyxl 并安装。")
        return

    proj_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    xlsx_path = os.path.join(proj_root, '附件', 'result3.xlsx')
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb = None
    if os.path.exists(xlsx_path) and load_workbook is not None:
        try:
            wb = load_workbook(xlsx_path)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    # 准备 / 复位工作表
    if 'Q5' in wb.sheetnames:
        ws = wb['Q5']
        wb.remove(ws)
        ws = wb.create_sheet('Q5')
    else:
        ws = wb.create_sheet('Q5')

    # 表头
    ws.append(['UAV', 'Bomb', '目标导弹', '航向角(度)', '速度(m/s)', '投放时间(s)', '引信时间(s)',
               '投放点x(m)', '投放点y(m)', '投放点z(m)',
               '起爆点x(m)', '起爆点y(m)', '起爆点z(m)'])

    for bomb in bombs_data_precise:
        pr = bomb['P_rel']; e = bomb['E']
        i = bomb['uav_idx']; k = bomb['bomb_idx']
        angle, speed, t_r1, dt_r2, dt_r3, t_f_list = uavs[i]
        if k == 0:
            t_rel = t_r1; t_fuz = t_f_list[0]
        elif k == 1:
            t_rel = t_r1 + dt_r2; t_fuz = t_f_list[1]
        else:
            t_rel = t_r1 + dt_r2 + dt_r3; t_fuz = t_f_list[2]
        ws.append([
            f'FY{i+1}', f'B{k+1}', f'M{bomb["m_idx"]+1}',
            round(angle, 3), round(speed, 3), round(t_rel, 3), round(t_fuz, 3),
            round(pr[0], 3), round(pr[1], 3), round(pr[2], 3),
            round(e[0], 3), round(e[1], 3), round(e[2], 3)
        ])

    # 总结
    ws.append(['M1并集时长(s)', round(per_missile[0], 3)])
    ws.append(['M2并集时长(s)', round(per_missile[1], 3)])
    ws.append(['M3并集时长(s)', round(per_missile[2], 3)])
    ws.append(['总并集时长和(s)', round(total_duration_precise, 3)])

    try:
        wb.save(xlsx_path)
        print("Excel 已写入:", xlsx_path)
    except Exception as e:
        print("保存 Excel 失败:", e)


if __name__ == "__main__":
    best_params, best_fit = run_pso_optimization()
    if best_params:
        format_and_output_results(best_params)


